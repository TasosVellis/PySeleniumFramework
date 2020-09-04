import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname": "Tasos", "lastname": "Vellis", "gender": "Male"}, {"firstname": "Angie", "lastname": "Grigoropoulou", "gender": "Female"}]

    @staticmethod
    def getTestData(test_case_name):
        testdata = {}
        book = openpyxl.load_workbook("C:\\Users\\Angie\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # dict ["lastname']="Vellis
                    testdata[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [testdata]
